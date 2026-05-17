"""Tests cho module excel_export.py — 25 test cases.

Covers: AC #1-12, edge cases, styling, confidence conversion, Vietnamese diacritics.
"""

import os
import tempfile
import unittest
from datetime import datetime
from unittest.mock import patch, MagicMock

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from src.core.excel_export import ExcelExporter


def _make_session_result(present=None, absent=None, start_time=None, end_time=None, class_id=1):
    """Helper: tạo mock session_result dict."""
    return {
        'class_id': class_id,
        'start_time': start_time or datetime(2026, 4, 15, 8, 0, 0),
        'end_time': end_time or datetime(2026, 4, 15, 9, 0, 0),
        'present': present if present is not None else [
            {
                'id': 1,
                'student_code': '2024002',
                'name': 'Trần Văn B',
                'is_present': True,
                'confidence': 0.95,
                'image_path': '/tmp/img1.jpg',
                'mark_time': datetime(2026, 4, 15, 8, 15, 30),
            },
        ],
        'absent': absent if absent is not None else [
            {
                'id': 2,
                'student_code': '2024001',
                'name': 'Nguyễn Văn A',
                'is_present': False,
                'confidence': 0.0,
                'image_path': None,
                'mark_time': None,
            },
        ],
    }


def _make_class_info():
    return {'id': 1, 'class_code': 'CS101-01', 'subject': 'Lập trình Python', 'teacher_id': 1}


def _make_teacher_info():
    return {'id': 1, 'name': 'TS. Nguyễn Văn X', 'teacher_code': 'GV001', 'photo_path': '', 'added_date': ''}


class TestExcelExporter(unittest.TestCase):
    """Test suite cho ExcelExporter — 25 tests."""

    def setUp(self):
        """Patch Config Singleton để redirect exports_dir tới tmp."""
        # F3-fix: use tempfile.mkdtemp() instead of repo-relative dir
        self.tmp_dir = tempfile.mkdtemp(prefix='test_excel_exports_')

        patcher = patch('src.core.excel_export.Config')
        self.mock_config_cls = patcher.start()
        self.addCleanup(patcher.stop)
        mock_config_instance = MagicMock()
        mock_config_instance.get.return_value = self.tmp_dir
        self.mock_config_cls.return_value = mock_config_instance

        self.exporter = ExcelExporter()

    def tearDown(self):
        """Dọn file test."""
        import shutil
        if os.path.exists(self.tmp_dir):
            shutil.rmtree(self.tmp_dir)

    # ── Test 1: File được tạo đúng tên (AC#5) ──
    def test_file_created_with_correct_name(self):
        session = _make_session_result()
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        filename = os.path.basename(result_path)
        self.assertEqual(filename, '2026-04-15_CS101-01.xlsx')

    # ── Test 2: File được lưu đúng thư mục (AC#5) ──
    def test_file_saved_in_exports_dir(self):
        session = _make_session_result()
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        self.assertTrue(result_path.startswith(os.path.abspath(self.tmp_dir)))
        self.assertTrue(os.path.exists(result_path))

    # ── Test 3: Header metadata (AC#4) ──
    def test_header_metadata_rows(self):
        session = _make_session_result()
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        wb = load_workbook(result_path)
        ws = wb.active

        self.assertEqual(ws.cell(row=1, column=1).value, "Ngày:")
        self.assertEqual(ws.cell(row=1, column=2).value, "15/04/2026")
        self.assertEqual(ws.cell(row=2, column=1).value, "Môn học:")
        self.assertEqual(ws.cell(row=2, column=2).value, "Lập trình Python")
        self.assertEqual(ws.cell(row=3, column=1).value, "Mã lớp:")
        self.assertEqual(ws.cell(row=3, column=2).value, "CS101-01")
        self.assertEqual(ws.cell(row=4, column=1).value, "Giảng viên:")
        self.assertEqual(ws.cell(row=4, column=2).value, "TS. Nguyễn Văn X")
        self.assertEqual(ws.cell(row=5, column=1).value, "Thời gian:")
        self.assertEqual(ws.cell(row=5, column=2).value, "08:00 - 09:00")
        wb.close()

    # ── Test 4: Header bảng (AC#3) ──
    def test_table_header_row(self):
        session = _make_session_result()
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        wb = load_workbook(result_path)
        ws = wb.active

        expected = ['STT', 'MSSV', 'Họ và tên', 'Trạng thái', 'Giờ điểm danh', 'Độ khớp (%)']
        for col_idx, header in enumerate(expected, start=1):
            self.assertEqual(ws.cell(row=7, column=col_idx).value, header)
        wb.close()

    # ── Test 5: Present student data (AC#3) ──
    def test_present_student_data(self):
        session = _make_session_result()
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        wb = load_workbook(result_path)
        ws = wb.active

        # After sort by MSSV: 2024001 (absent) row 8, 2024002 (present) row 9
        present_row = 9
        self.assertEqual(ws.cell(row=present_row, column=4).value, "Có mặt")
        self.assertEqual(ws.cell(row=present_row, column=5).value, "08:15:30")
        self.assertEqual(ws.cell(row=present_row, column=6).value, "95.0%")
        wb.close()

    # ── Test 6: Absent student data (AC#3) ──
    def test_absent_student_data(self):
        session = _make_session_result()
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        wb = load_workbook(result_path)
        ws = wb.active

        # After sort by MSSV: 2024001 (absent) row 8
        absent_row = 8
        self.assertEqual(ws.cell(row=absent_row, column=4).value, "Vắng")
        self.assertIn(ws.cell(row=absent_row, column=5).value, ("", None))
        self.assertIn(ws.cell(row=absent_row, column=6).value, ("", None))
        wb.close()

    # ── Test 7: Sắp xếp theo MSSV tăng dần (AC#7) ──
    def test_sorted_by_student_code(self):
        session = _make_session_result()
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        wb = load_workbook(result_path)
        ws = wb.active

        # 2024001 (Nguyễn Văn A) should come before 2024002 (Trần Văn B)
        self.assertEqual(ws.cell(row=8, column=2).value, '2024001')
        self.assertEqual(ws.cell(row=9, column=2).value, '2024002')
        wb.close()

    # ── Test 8: Summary row (AC#8) ──
    def test_summary_row(self):
        session = _make_session_result()
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        wb = load_workbook(result_path)
        ws = wb.active

        # 2 students total, 1 present → summary at row 10
        summary_cell = ws.cell(row=10, column=1)
        self.assertEqual(summary_cell.value, "Tổng kết: Có mặt 1/2 (50.0%)")
        self.assertTrue(summary_cell.font.bold)
        wb.close()

    # ── Test 9: Styling — table header (AC#6) ──
    def test_table_header_styling(self):
        session = _make_session_result()
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        wb = load_workbook(result_path)
        ws = wb.active

        for col_idx in range(1, 7):
            cell = ws.cell(row=7, column=col_idx)
            self.assertTrue(cell.font.bold)
            self.assertEqual(cell.fill.start_color.rgb, '004472C4')  # openpyxl prefixes with '00'
        wb.close()

    # ── Test 10: Edge case — no students (AC#11) ──
    def test_no_students_empty_session(self):
        session = _make_session_result(present=[], absent=[])
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        wb = load_workbook(result_path)
        ws = wb.active

        # Summary at row 8 (7 + 0 + 1)
        summary_cell = ws.cell(row=8, column=1)
        self.assertEqual(summary_cell.value, "Tổng kết: Có mặt 0/0 (0.0%)")
        wb.close()

    # ── Test 11: session_result None → ValueError (AC#9) ──
    def test_session_result_none_raises_value_error(self):
        with self.assertRaises(ValueError) as ctx:
            self.exporter.export_session(None, _make_class_info(), _make_teacher_info())
        self.assertIn("rỗng hoặc None", str(ctx.exception))

    # ── Test 12: Present SV có mark_time=None → defensive (AC#11) ──
    def test_present_student_with_none_mark_time(self):
        present = [{
            'id': 1, 'student_code': '2024001', 'name': 'Test',
            'is_present': True, 'confidence': 0.85,
            'image_path': None, 'mark_time': None,
        }]
        session = _make_session_result(present=present, absent=[])
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        wb = load_workbook(result_path)
        ws = wb.active

        # mark_time=None → empty string / None in openpyxl
        self.assertIn(ws.cell(row=8, column=5).value, ("", None))
        # confidence=0.85 → "85.0%"
        self.assertEqual(ws.cell(row=8, column=6).value, "85.0%")
        wb.close()

    # ── Test 13: Return value is absolute path (AC#12) ──
    def test_return_value_is_absolute_path(self):
        session = _make_session_result()
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        self.assertTrue(os.path.isabs(result_path))
        self.assertTrue(os.path.exists(result_path))

    # ── Test 14: Auto tạo thư mục exports (AC#5) ──
    def test_auto_create_exports_dir(self):
        # Remove and recreate exporter
        import shutil
        nested_dir = os.path.join(self.tmp_dir, 'nested', 'exports')
        if os.path.exists(nested_dir):
            shutil.rmtree(nested_dir)

        mock_config_instance = MagicMock()
        mock_config_instance.get.return_value = nested_dir
        self.mock_config_cls.return_value = mock_config_instance

        exporter = ExcelExporter()
        self.assertTrue(os.path.exists(nested_dir))

    # ── Test 15: Tên file trùng → ghi đè (AC#11) ──
    def test_overwrite_existing_file(self):
        session = _make_session_result()
        path1 = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        path2 = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        self.assertEqual(path1, path2)
        self.assertTrue(os.path.exists(path2))

    # ── Test 16: Vietnamese diacritics (AC#3) ──
    def test_vietnamese_diacritics(self):
        present = [{
            'id': 1, 'student_code': '2024001', 'name': 'Nguyễn Thị Ái',
            'is_present': True, 'confidence': 0.9,
            'image_path': None, 'mark_time': datetime(2026, 4, 15, 8, 20),
        }]
        session = _make_session_result(present=present, absent=[])
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        wb = load_workbook(result_path)
        ws = wb.active
        self.assertEqual(ws.cell(row=8, column=3).value, 'Nguyễn Thị Ái')
        wb.close()

    # ── Test 17: class_info None → defaults (AC#11) ──
    def test_class_info_none_uses_defaults(self):
        session = _make_session_result()
        result_path = self.exporter.export_session(session, None, _make_teacher_info())
        wb = load_workbook(result_path)
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=2).value, "N/A")   # subject
        self.assertEqual(ws.cell(row=3, column=2).value, "N/A")   # class_code
        wb.close()

    # ── Test 18: teacher_info None → defaults (AC#11) ──
    def test_teacher_info_none_uses_defaults(self):
        session = _make_session_result()
        result_path = self.exporter.export_session(session, _make_class_info(), None)
        wb = load_workbook(result_path)
        ws = wb.active
        self.assertEqual(ws.cell(row=4, column=2).value, "N/A")   # teacher name
        wb.close()

    # ── Test 19: session_result empty dict → ValueError (AC#9) ──
    def test_session_result_empty_dict_raises_value_error(self):
        with self.assertRaises(ValueError):
            self.exporter.export_session({}, _make_class_info(), _make_teacher_info())

    # ── Test 20: wb.save() PermissionError → re-raise as OSError (AC#9) ──
    def test_save_permission_error_raises_os_error(self):
        session = _make_session_result()
        with patch('src.core.excel_export.Workbook') as mock_wb_cls:
            mock_wb = MagicMock()
            mock_ws = MagicMock()
            mock_wb.active = mock_ws
            mock_ws.cell.return_value = MagicMock()
            mock_wb.save.side_effect = PermissionError("Permission denied")
            mock_wb_cls.return_value = mock_wb

            with self.assertRaises(OSError) as ctx:
                self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
            self.assertIn("Không thể ghi file Excel", str(ctx.exception))

    # ── Test 21: Confidence conversion (AC — CRITICAL) ──
    def test_confidence_conversion_decimal_to_percent(self):
        present = [{
            'id': 1, 'student_code': '2024001', 'name': 'Test',
            'is_present': True, 'confidence': 0.85,
            'image_path': None, 'mark_time': datetime(2026, 4, 15, 8, 10),
        }]
        session = _make_session_result(present=present, absent=[])
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        wb = load_workbook(result_path)
        ws = wb.active
        # 0.85 * 100 = 85.0%  — NOT "0.9%"
        self.assertEqual(ws.cell(row=8, column=6).value, "85.0%")
        wb.close()

    # ── Test 22: start_time=None → uses fallback datetime (F1/F6) ──
    def test_start_time_none_uses_fallback(self):
        session = _make_session_result()
        session['start_time'] = None
        session['end_time'] = None
        # Should not crash — uses datetime.now() fallback
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        self.assertTrue(os.path.exists(result_path))
        wb = load_workbook(result_path)
        ws = wb.active
        # Row 1 should have "Ngày:" with a valid date string
        self.assertEqual(ws.cell(row=1, column=1).value, "Ngày:")
        self.assertIsNotNone(ws.cell(row=1, column=2).value)
        wb.close()

    # ── Test 23: Present student with confidence=None → safe (F4) ──
    def test_present_student_confidence_none_safe(self):
        present = [{
            'id': 1, 'student_code': '2024001', 'name': 'Test',
            'is_present': True, 'confidence': None,
            'image_path': None, 'mark_time': datetime(2026, 4, 15, 8, 10),
        }]
        session = _make_session_result(present=present, absent=[])
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        wb = load_workbook(result_path)
        ws = wb.active
        # confidence=None → empty display
        self.assertIn(ws.cell(row=8, column=6).value, ("", None))
        wb.close()

    # ── Test 24: Malformed student data with missing keys → graceful (F4) ──
    def test_malformed_student_missing_keys(self):
        """Student dict missing standard keys should use .get() defaults, not crash."""
        present = [{'id': 1, 'is_present': True}]  # missing student_code, name, confidence, mark_time
        absent = [{'id': 2}]  # missing everything except id
        session = _make_session_result(present=present, absent=absent)
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        self.assertTrue(os.path.exists(result_path))
        wb = load_workbook(result_path)
        ws = wb.active
        # Should have 2 data rows (8, 9) + summary at row 10
        summary_cell = ws.cell(row=10, column=1)
        self.assertIn("Tổng kết", summary_cell.value)
        wb.close()

    # ── Test 25: None entries in student lists → filtered out (F5) ──
    def test_none_entries_in_student_lists_filtered(self):
        """None entries mixed into present/absent lists should be silently filtered."""
        present = [
            None,
            {
                'id': 1, 'student_code': '2024001', 'name': 'Valid Student',
                'is_present': True, 'confidence': 0.9,
                'image_path': None, 'mark_time': datetime(2026, 4, 15, 8, 20),
            },
            None,
        ]
        absent = [None]
        session = _make_session_result(present=present, absent=absent)
        result_path = self.exporter.export_session(session, _make_class_info(), _make_teacher_info())
        self.assertTrue(os.path.exists(result_path))
        wb = load_workbook(result_path)
        ws = wb.active
        # Only 1 valid student should appear
        self.assertEqual(ws.cell(row=8, column=2).value, '2024001')
        summary_cell = ws.cell(row=9, column=1)
        self.assertEqual(summary_cell.value, "Tổng kết: Có mặt 1/1 (100.0%)")
        wb.close()


if __name__ == '__main__':
    unittest.main()
