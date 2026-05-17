"""Module xuất kết quả điểm danh ra file Excel.

Sử dụng openpyxl để tạo file .xlsx với:
- Header metadata (Ngày, Môn, Lớp, GV, Thời gian)
- Bảng điểm danh (STT, MSSV, Họ tên, Trạng thái, Giờ, % khớp)
- Summary row (Tổng kết: Có mặt X/Y (Z%))
"""

import logging
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.core.config import Config

logger = logging.getLogger(__name__)

# ── Styling Constants ──────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
METADATA_FONT = Font(bold=True, size=11)
DATA_FONT = Font(size=11)
SUMMARY_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')

# Column widths (approximate minimum)
COLUMN_WIDTHS = {'A': 6, 'B': 15, 'C': 25, 'D': 12, 'E': 15, 'F': 12}

# Table header labels
TABLE_HEADERS = ['STT', 'MSSV', 'Họ và tên', 'Trạng thái', 'Giờ điểm danh', 'Độ khớp (%)']


class ExcelExporter:
    """Xuất kết quả điểm danh ra file Excel (.xlsx).

    Thuộc Model layer (src/core/). Không import từ src/gui/.
    Nhận data qua parameters từ Presenter — KHÔNG tự query database.
    """

    def __init__(self):
        config = Config()
        self.exports_dir = config.get('paths', 'exports_dir', default='data/exports')
        os.makedirs(self.exports_dir, exist_ok=True)

    def export_session(self, session_result: dict, class_info: dict = None, teacher_info: dict = None) -> str:
        """Xuất kết quả điểm danh ra file Excel.

        Args:
            session_result: Dict từ AttendanceSession.end_session()
            class_info: Dict từ DatabaseManager.get_class() (None → dùng defaults)
            teacher_info: Dict từ DatabaseManager.get_teacher() (None → dùng defaults)

        Returns:
            str: Absolute path của file Excel đã tạo

        Raises:
            ValueError: Nếu session_result rỗng/None/empty dict
            OSError: Nếu không thể ghi file
        """
        # ── Validate input (AC#9: catches None AND {}) ──
        if not session_result:
            logger.warning("export_session() called with invalid session_result: %s", type(session_result))
            raise ValueError("Dữ liệu phiên điểm danh không hợp lệ (rỗng hoặc None).")

        logger.info("Bắt đầu xuất Excel cho session...")

        # ── Safe key access with defaults (F2) ──
        # F2-fix: explicit `or []` guards against None values stored in dict
        present_list = session_result.get('present') or []
        absent_list = session_result.get('absent') or []
        # F1-fix: .get() returns None if key exists with None value — guard with `or`
        start_time = session_result.get('start_time') or datetime.now()
        end_time = session_result.get('end_time') or datetime.now()

        # ── Null-safe defaults (AC#11) ──
        _class_info = class_info or {}
        _teacher_info = teacher_info or {}
        class_code = _class_info.get('class_code', 'N/A')
        subject = _class_info.get('subject', 'N/A')
        teacher_name = _teacher_info.get('name', 'N/A')

        # ── Create workbook ──
        wb = Workbook()
        ws = wb.active
        ws.title = "Điểm danh"

        # ── Header metadata (rows 1-5) — AC#4 ──
        metadata_rows = [
            ("Ngày:", start_time.strftime('%d/%m/%Y')),
            ("Môn học:", subject),
            ("Mã lớp:", class_code),
            ("Giảng viên:", teacher_name),
            ("Thời gian:", f"{start_time.strftime('%H:%M')} - {end_time.strftime('%H:%M')}"),
        ]
        for row_idx, (label, value) in enumerate(metadata_rows, start=1):
            cell_label = ws.cell(row=row_idx, column=1, value=label)
            cell_value = ws.cell(row=row_idx, column=2, value=value)
            self._apply_header_style(cell_label)
            self._apply_header_style(cell_value)

        # ── Table header (row 7) — AC#3, AC#6 ──
        for col_idx, header in enumerate(TABLE_HEADERS, start=1):
            cell = ws.cell(row=7, column=col_idx, value=header)
            self._apply_table_header_style(cell)

        # ── Merge + sort data — AC#7 ──
        # F5-fix: filter out None entries from student lists (defensive)
        all_students = [s for s in list(present_list) + list(absent_list) if s is not None]
        # F1-fix: str() cast prevents TypeError on None/int student_code
        all_students.sort(key=lambda s: str(s.get('student_code', '')))

        # ── Data rows (row 8+) — AC#3 ──
        present_count = 0
        for idx, student in enumerate(all_students, start=1):
            row_num = 7 + idx
            is_present = student.get('is_present', False)

            if is_present:
                present_count += 1
                status = "Có mặt"
                mark_time = student.get('mark_time')
                time_str = mark_time.strftime('%H:%M:%S') if mark_time else ""
                confidence = student.get('confidence')
                # CRITICAL: confidence lưu dạng decimal 0.0-1.0, phải nhân 100
                # F4-fix: explicit None check — truthiness alone is fragile
                if confidence is not None and confidence > 0:
                    confidence_str = f"{confidence * 100:.1f}%"
                else:
                    confidence_str = ""
            else:
                status = "Vắng"
                time_str = ""
                confidence_str = ""

            data = [
                idx,                                    # STT
                student.get('student_code', ''),        # MSSV
                student.get('name', ''),                # Họ và tên
                status,                                 # Trạng thái
                time_str,                               # Giờ điểm danh
                confidence_str,                         # Độ khớp (%)
            ]
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                self._apply_data_style(cell)
                # F8-fix: Name column (C) uses LEFT alignment for readability
                if col_idx == 3:
                    cell.alignment = LEFT_ALIGN

        # ── Summary row — AC#8 ──
        total_count = len(all_students)
        percent = (present_count / total_count * 100) if total_count > 0 else 0
        summary_row = 7 + len(all_students) + 1
        summary_text = f"Tổng kết: Có mặt {present_count}/{total_count} ({percent:.1f}%)"

        ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=6)
        summary_cell = ws.cell(row=summary_row, column=1, value=summary_text)
        self._apply_summary_style(summary_cell)

        # ── Column widths — AC#6 ──
        for col_letter, width in COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

        # ── Generate filename — AC#5 ──
        date_str = start_time.strftime('%Y-%m-%d')
        safe_class_code = class_code.replace('/', '-').replace('\\', '-')
        filename = f"{date_str}_{safe_class_code}.xlsx"
        filepath = os.path.join(self.exports_dir, filename)

        # ── Save workbook — AC#9 ──
        try:
            wb.save(filepath)
        except OSError as e:
            logger.error(f"Không thể ghi file Excel: {e}")
            raise OSError(f"Không thể ghi file Excel: {e}") from e  # F5-fix: preserve chain
        finally:
            wb.close()  # F2-fix: release in-memory XML trees

        abs_path = os.path.abspath(filepath)
        logger.info(f"Đã xuất file Excel: {abs_path}")
        return abs_path

    # ── Helper methods ──

    @staticmethod
    def _apply_header_style(cell):
        """Áp dụng style cho header metadata (bold)."""
        cell.font = METADATA_FONT

    @staticmethod
    def _apply_table_header_style(cell):
        """Áp dụng style cho header bảng (bold + fill + white font + border)."""
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    @staticmethod
    def _apply_data_style(cell):
        """Áp dụng style cho data rows (thin border + center alignment)."""
        cell.font = DATA_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    @staticmethod
    def _apply_summary_style(cell):
        """Áp dụng style cho summary row (bold)."""
        cell.font = SUMMARY_FONT
